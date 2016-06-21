from openpyxl import Workbook
import sys
from io import BytesIO

from . import scriptDB

def xlsfile(inputdata, inputcookie):
	out = scriptDB.readQuery(inputdata['idquery'])
	if out is None  :
		print(id + " no encontrado")
		return None
	user = scriptDB.unfold(inputcookie)
	folder = scriptDB.getFolder({"idquery": out["folder"]}, inputcookie, False)
	
	operations = [{"$match": {} }]

	if inputdata["typeQuery"] is "1": #aprobado por mi
		operations[0]["$match"]["$and"] = [{"results.vote.yes": str(user["_id"])}]
	elif inputdata["typeQuery"] is "2": #aprobado por  todos
		aux = []
		for userId in folder["user"] :
			aux.append( {"results.vote.yes": userId} )
		operations[0]["$match"]["$and"] = aux
	
	elif inputdata["typeQuery"] is "3": #rechazados por mi
		operations[0]["$match"]["$and"] = [{"results.vote.no": str(user["_id"])}]
	elif inputdata["typeQuery"] is "4": #rechazado por todos
		aux = []
		for userId in folder["user"] :
			aux.append( {"results.vote.no": userId} )
		operations[0]["$match"]["$and"] = aux

	elif inputdata["typeQuery"] is "5": #no votado por mi
		operations[0]["$match"]["$and"] = [{"results.vote.yes": {"$ne": str(user["_id"]) }}, {"results.vote.no": {"$ne": str(user["_id"]) }}]
	elif inputdata["typeQuery"] is "6": #no votado por todos
		aux = []
		for userId in folder["user"] : 
			aux.extend( [{"results.vote.yes": {"$ne": userId }}, {"results.vote.no": {"$ne": userId }}] )
		operations[0]["$match"]["$and"] = aux
	

	elif inputdata["typeQuery"] is "7": #en discusión
		operations[0]["$match"]["$or"] = [{"results.vote.yes": { "$exists": True } }, {"results.vote.no": { "$exists": True } }]
		operations.append({ "$project": {
			"results": "$results", 
			"yes": {"$size": { "$ifNull": [ "$results.vote.yes", [] ] }}, 
			"no": {"$size": { "$ifNull": [ "$results.vote.no", [] ] }}
			}})
		operations.append({ "$match": {
			"$or": [{"yes": { "$gt": 0, "$lt": len(folder["user"]) } }, {"no": { "$gt": 0, "$lt": len(folder["user"]) } }]
			}})


	if (inputdata.get("duplicate") is None) or (inputdata.get("duplicate") is "off") : #si no se piden duplicados
		if operations[0]["$match"].get("$and") is None:
			operations[0]["$match"]["$and"] = []
		operations[0]["$match"]["$and"].append({"results.isDuplicate": {"$in": [False, None]} })	

	operations.append( { "$group": {"_id": '$_id', "results": {"$push": "$results"}} } )


	#try:
	wb = Workbook()
	ws = wb.active
	ws.title = "resumen"
	ws['A'+ str(1)] = "query" 
	ws['B'+ str(1)] = out.get("query")
	ws['A'+ str(2)] = "date" 
	ws['B'+ str(2)] = out.get("date")

	for doc in out["sources"]:
		ws = wb.create_sheet()
		ws.title = doc["name"]
		ws['A'+ str(1)] = "rank"
		ws['B'+ str(1)] = "title"
		ws['C'+ str(1)] = "authors"
		ws['D'+ str(1)] = "abstract"
		ws['E'+ str(1)] = "url"
		ws['F'+ str(1)] = "publication name"
		ws['G'+ str(1)] = "publication date"
		ws['H'+ str(1)] = "publication page"
		ws['I'+ str(1)] = "doi"
		ws['I'+ str(1)] = "votes accept"
		ws['J'+ str(1)] = "votes reject"
		ws['K'+ str(1)] = "comments"
		row = 2
		docAux = scriptDB.simpleAggregateSource(doc, operations=operations)
		if docAux:
			for item in docAux['results']:
				ws['A'+ str(row)] = item.get('rank')
				ws['B'+ str(row)] = item.get('title')
				ws['C'+ str(row)] = item.get('authors')
				ws['D'+ str(row)] = item.get('abstract')
				ws['E'+ str(row)] = item.get('mdurl')
				ws['F'+ str(row)] = item.get('pubN')
				ws['G'+ str(row)] = item.get('pubY')
				ws['H'+ str(row)] = item.get('pubP')
				ws['I'+ str(row)] = item.get('doi')
				ws['I'+ str(row)] = 0
				ws['J'+ str(row)] = 0
				ws['K'+ str(row)] = 0
				if item["vote"].get('yes'):
					ws['I'+ str(row)] = len(item["vote"].get('yes'))
				if item["vote"].get('no'):
					ws['J'+ str(row)] = len(item["vote"].get('no'))
				if item.get('comment'):
					ws['K'+ str(row)] = len(item.get('comment'))
				row += 1
	print("saving " + str(out["_id"]) + '.xlsx')
	out = BytesIO()
	wb.save(out)
	print("saved")
	return out
	#except:
	#	print("error en xlsx")
	#	return None