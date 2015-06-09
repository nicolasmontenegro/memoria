from pymongo import MongoClient
import time
from bson.objectid import ObjectId
from openpyxl import Workbook
import sys


def xlsfile(listres, id):
	try:
		wb = Workbook()
		for doc in listres:
			ws = wb.create_sheet()
			ws.title = doc['name'][:10]
			ws['A'+ str(1)] = "rank"
			ws['B'+ str(1)] = "title"
			ws['C'+ str(1)] = "authors"
			ws['D'+ str(1)] = "abstract"
			ws['E'+ str(1)] = "url"
			ws['F'+ str(1)] = "publication name"
			ws['G'+ str(1)] = "publication date"
			ws['H'+ str(1)] = "publication page"
			ws['I'+ str(1)] = "doi"
			row = 2
			for item in doc['data']['res']:
				ws['A'+ str(row)] = item['rank']
				ws['B'+ str(row)] = item['title']
				ws['C'+ str(row)] = item['authors']
				ws['D'+ str(row)] = item['abstract']
				ws['E'+ str(row)] = item['mdurl']
				ws['F'+ str(row)] = item['pubN']
				ws['G'+ str(row)] = item['pubY']
				ws['H'+ str(row)] = item['pubP']
				ws['I'+ str(row)] = item['doi']
				row += 1
		strout = 'buscador/xls/' +  id + '.xlsx'
		wb.save(strout)
		return strout
	except:
		return ""
